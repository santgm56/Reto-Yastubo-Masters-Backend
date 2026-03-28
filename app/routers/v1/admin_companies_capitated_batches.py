from __future__ import annotations

import hashlib
import json
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.db.database import get_db
from app.services.auth_service import AuthService

router = APIRouter(prefix="/api/v1/admin/companies", tags=["admin-companies-capitated"])


def _extract_bearer_token(authorization: str | None) -> str:
    if not authorization:
        return ""

    raw = authorization.strip()
    if not raw.lower().startswith("bearer "):
        return ""

    return raw[7:].strip()


def _require_admin(request: Request, authorization: str | None, db: Session) -> dict:
    token = _extract_bearer_token(authorization)
    if not token:
        token = str(request.cookies.get("yastubo_access_token") or "").strip()

    if not token:
        raise HTTPException(status_code=403, detail="Forbidden")

    auth_payload = AuthService(db).me(token)
    role = str(auth_payload.get("role") or "").upper()
    if role != "ADMIN":
        raise HTTPException(status_code=403, detail="Forbidden")

    return auth_payload


def _require_company_exists(db: Session, company_id: int) -> None:
    row = db.execute(
        text(
            """
            SELECT id
            FROM companies
            WHERE id = :company_id
            LIMIT 1
            """
        ),
        {"company_id": int(company_id)},
    ).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="Not Found")


def _storage_root() -> Path:
    settings = get_settings()
    configured = Path(str(settings.frontend_storage_root or "").strip())
    if configured.is_absolute():
        return configured

    backend_root = Path(__file__).resolve().parents[3]
    return (backend_root / configured).resolve()


def _decode_product_name(raw_name: str | None) -> str:
    name_value = str(raw_name or "").strip()
    if name_value == "":
        return "Producto"

    if name_value.startswith("{") and name_value.endswith("}"):
        try:
            data = json.loads(name_value)
            if isinstance(data, dict):
                preferred = data.get("es") or data.get("en")
                if isinstance(preferred, str) and preferred.strip():
                    return preferred.strip()
        except (TypeError, ValueError):
            return name_value

    return name_value


def _excel_sheet_title(base_title: str, used_titles: set[str]) -> str:
    cleaned = " ".join(str(base_title).replace("\\", " ").replace("/", " ").replace("?", " ").replace("*", " ").replace("[", " ").replace("]", " ").replace(":", " ").split())
    if cleaned == "":
        cleaned = "Hoja"

    cleaned = cleaned[:31]
    candidate = cleaned
    suffix_index = 2

    while candidate in used_titles:
        suffix = f" ({suffix_index})"
        candidate = f"{cleaned[: max(0, 31 - len(suffix))]}{suffix}"
        suffix_index += 1

    used_titles.add(candidate)
    return candidate


def _normalize_coverage_month(raw_coverage_month: str | None) -> date:
    if not raw_coverage_month:
        now = datetime.utcnow()
        return date(now.year, now.month, 1)

    try:
        parsed = datetime.fromisoformat(str(raw_coverage_month).strip())
    except ValueError as exc:
        raise HTTPException(status_code=422, detail="Invalid coverage_month") from exc

    return date(parsed.year, parsed.month, 1)


def _serialize_batch_row(row) -> dict:
    created_by_payload = None
    created_by_id = row.get("created_by_id")
    if created_by_id is not None:
        display_name = str(row.get("created_by_display_name") or "").strip()
        if not display_name:
            first_name = str(row.get("created_by_first_name") or "").strip()
            last_name = str(row.get("created_by_last_name") or "").strip()
            display_name = f"{first_name} {last_name}".strip()
        created_by_payload = {
            "id": int(created_by_id),
            "display_name": display_name,
            "email": row.get("created_by_email"),
        }

    return {
        "id": int(row.get("id") or 0),
        "company_id": int(row.get("company_id") or 0),
        "coverage_month": row.get("coverage_month"),
        "source": row.get("source"),
        "source_file_id": row.get("source_file_id"),
        "original_filename": row.get("original_filename"),
        "file_hash": row.get("file_hash"),
        "created_by_user_id": row.get("created_by_user_id"),
        "created_by": created_by_payload,
        "status": row.get("status"),
        "processed_at": row.get("processed_at"),
        "rolled_back_at": row.get("rolled_back_at"),
        "rolled_back_by_user_id": row.get("rolled_back_by_user_id"),
        "total_rows": int(row.get("total_rows") or 0),
        "total_applied": int(row.get("total_applied") or 0),
        "total_rejected": int(row.get("total_rejected") or 0),
        "total_duplicated": int(row.get("total_duplicated") or 0),
        "total_incongruences": int(row.get("total_incongruences") or 0),
        "total_plan_errors": int(row.get("total_plan_errors") or 0),
        "total_rolled_back": int(row.get("total_rolled_back") or 0),
        "is_any_month_allowed": bool(row.get("is_any_month_allowed") or 0),
        "cutoff_day": row.get("cutoff_day"),
        "error_summary": row.get("error_summary"),
        "summary_json": row.get("summary_json"),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
        "file_temporary_url": f"/api/v1/files/{row.get('source_file_uuid')}" if row.get("source_file_uuid") else None,
    }


def _can_rollback_batch(batch_row) -> bool:
    status = str(batch_row.get("status") or "").lower()
    if status not in {"processed", "processed_zero"}:
        return False
    if batch_row.get("rolled_back_at") is not None:
        return False
    return True


def _can_rollback_monthly_record(monthly_row) -> bool:
    status = str(monthly_row.get("status") or "").lower()
    return status == "active"


def _decode_product_name_for_response(raw_name: str | None) -> str:
    return _decode_product_name(raw_name)


@router.get("/{company_id}/capitated/batches/{batch_id:int}")
def batches_show(
    company_id: int,
    batch_id: int,
    request: Request,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
) -> dict:
    _require_admin(request=request, authorization=authorization, db=db)
    _require_company_exists(db=db, company_id=int(company_id))

    row = db.execute(
        text(
            """
            SELECT
                b.id,
                b.company_id,
                b.coverage_month,
                b.source,
                b.source_file_id,
                b.original_filename,
                b.file_hash,
                b.created_by_user_id,
                b.status,
                b.processed_at,
                b.rolled_back_at,
                b.rolled_back_by_user_id,
                b.total_rows,
                b.total_applied,
                b.total_rejected,
                b.total_duplicated,
                b.total_incongruences,
                b.total_plan_errors,
                b.total_rolled_back,
                b.is_any_month_allowed,
                b.cutoff_day,
                b.error_summary,
                b.summary_json,
                b.created_at,
                b.updated_at,
                f.uuid AS source_file_uuid,
                u.id AS created_by_id,
                u.display_name AS created_by_display_name,
                u.first_name AS created_by_first_name,
                u.last_name AS created_by_last_name,
                u.email AS created_by_email
            FROM capitados_batch_logs b
            LEFT JOIN files f ON f.id = b.source_file_id
            LEFT JOIN users u ON u.id = b.created_by_user_id
            WHERE b.id = :batch_id
              AND b.company_id = :company_id
            LIMIT 1
            """
        ),
        {"batch_id": int(batch_id), "company_id": int(company_id)},
    ).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="Not Found")

    payload = _serialize_batch_row(row)
    payload["can_rollback"] = _can_rollback_batch(row)

    return {"batch": payload}


@router.get("/{company_id}/capitated/batches/{batch_id:int}/items")
def batches_items(
    company_id: int,
    batch_id: int,
    request: Request,
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=25, ge=1, le=250),
    result: str | None = Query(default=None),
    sheet: str | None = Query(default=None),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
) -> dict:
    _require_admin(request=request, authorization=authorization, db=db)
    _require_company_exists(db=db, company_id=int(company_id))

    batch_exists = db.execute(
        text(
            """
            SELECT id
            FROM capitados_batch_logs
            WHERE id = :batch_id
              AND company_id = :company_id
            LIMIT 1
            """
        ),
        {"batch_id": int(batch_id), "company_id": int(company_id)},
    ).mappings().first()
    if not batch_exists:
        raise HTTPException(status_code=404, detail="Not Found")

    sheets_rows = db.execute(
        text(
            """
            SELECT DISTINCT sheet_name
            FROM capitados_batch_item_logs
            WHERE batch_id = :batch_id
            ORDER BY sheet_name ASC
            """
        ),
        {"batch_id": int(batch_id)},
    ).mappings().all()
    sheets = [str(row.get("sheet_name") or "") for row in sheets_rows if row.get("sheet_name") is not None]

    selected_sheet = None
    requested_sheet = str(sheet or "").strip()
    if requested_sheet and requested_sheet in sheets:
        selected_sheet = requested_sheet
    elif sheets:
        selected_sheet = sheets[0]

    where_clauses = ["i.batch_id = :batch_id"]
    params: dict[str, object] = {
        "batch_id": int(batch_id),
        "limit": int(per_page),
        "offset": int((page - 1) * per_page),
    }

    normalized_result = str(result or "").strip().lower()
    if normalized_result:
        where_clauses.append("LOWER(i.result) = :result")
        params["result"] = normalized_result

    if selected_sheet is not None:
        where_clauses.append("i.sheet_name = :sheet_name")
        params["sheet_name"] = selected_sheet

    where_sql = " AND ".join(where_clauses)

    total_row = db.execute(
        text(
            f"""
            SELECT COUNT(1) AS total
            FROM capitados_batch_item_logs i
            WHERE {where_sql}
            """
        ),
        params,
    ).mappings().first()
    total = int((total_row or {}).get("total") or 0)
    last_page = max(1, (total + per_page - 1) // per_page)

    rows = db.execute(
        text(
            f"""
            SELECT
                i.id,
                i.batch_id,
                i.sheet_name,
                i.row_number,
                i.product_id,
                i.plan_version_id,
                i.residence_raw,
                i.residence_code_extracted,
                i.repatriation_raw,
                i.repatriation_code_extracted,
                i.residence_country_id,
                i.repatriation_country_id,
                i.document_number,
                i.full_name,
                i.sex,
                i.age_reported,
                i.result,
                i.rejection_code,
                i.rejection_detail,
                i.person_id,
                i.contract_id,
                i.monthly_record_id,
                i.duplicated_record_id,
                i.created_at,
                i.updated_at,
                rc.id AS residence_country_ref_id,
                rc.name AS residence_country_name,
                rc.iso2 AS residence_country_iso2,
                rc.iso3 AS residence_country_iso3,
                rep.id AS repatriation_country_ref_id,
                rep.name AS repatriation_country_name,
                rep.iso2 AS repatriation_country_iso2,
                rep.iso3 AS repatriation_country_iso3
            FROM capitados_batch_item_logs i
            LEFT JOIN countries rc ON rc.id = i.residence_country_id
            LEFT JOIN countries rep ON rep.id = i.repatriation_country_id
            WHERE {where_sql}
            ORDER BY i.sheet_name ASC, i.row_number ASC
            LIMIT :limit OFFSET :offset
            """
        ),
        params,
    ).mappings().all()

    data = []
    for row in rows:
        data.append(
            {
                "id": int(row.get("id") or 0),
                "batch_id": int(row.get("batch_id") or 0),
                "sheet_name": row.get("sheet_name"),
                "row_number": row.get("row_number"),
                "product_id": row.get("product_id"),
                "plan_version_id": row.get("plan_version_id"),
                "residence_raw": row.get("residence_raw"),
                "residence_code_extracted": row.get("residence_code_extracted"),
                "repatriation_raw": row.get("repatriation_raw"),
                "repatriation_code_extracted": row.get("repatriation_code_extracted"),
                "residence_country_id": row.get("residence_country_id"),
                "repatriation_country_id": row.get("repatriation_country_id"),
                "document_number": row.get("document_number"),
                "full_name": row.get("full_name"),
                "sex": row.get("sex"),
                "age_reported": row.get("age_reported"),
                "result": row.get("result"),
                "rejection_code": row.get("rejection_code"),
                "rejection_detail": row.get("rejection_detail"),
                "person_id": row.get("person_id"),
                "contract_id": row.get("contract_id"),
                "monthly_record_id": row.get("monthly_record_id"),
                "duplicated_record_id": row.get("duplicated_record_id"),
                "created_at": row.get("created_at"),
                "updated_at": row.get("updated_at"),
                "residence_country": {
                    "id": row.get("residence_country_ref_id"),
                    "name": row.get("residence_country_name"),
                    "iso2": row.get("residence_country_iso2"),
                    "iso3": row.get("residence_country_iso3"),
                }
                if row.get("residence_country_ref_id") is not None
                else None,
                "repatriation_country": {
                    "id": row.get("repatriation_country_ref_id"),
                    "name": row.get("repatriation_country_name"),
                    "iso2": row.get("repatriation_country_iso2"),
                    "iso3": row.get("repatriation_country_iso3"),
                }
                if row.get("repatriation_country_ref_id") is not None
                else None,
            }
        )

    return {
        "data": data,
        "meta": {
            "current_page": int(page),
            "last_page": int(last_page),
            "per_page": int(per_page),
            "total": int(total),
        },
        "sheets": sheets,
        "filters": {
            "result": normalized_result if normalized_result else "",
            "sheet": selected_sheet,
        },
    }


@router.get("/{company_id}/capitated/batches/{batch_id:int}/monthly-records")
def batches_monthly_records(
    company_id: int,
    batch_id: int,
    request: Request,
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=25, ge=1, le=250),
    status: str | None = Query(default=None),
    product_id: int | None = Query(default=None),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
) -> dict:
    _require_admin(request=request, authorization=authorization, db=db)
    _require_company_exists(db=db, company_id=int(company_id))

    batch_exists = db.execute(
        text(
            """
            SELECT id
            FROM capitados_batch_logs
            WHERE id = :batch_id
              AND company_id = :company_id
            LIMIT 1
            """
        ),
        {"batch_id": int(batch_id), "company_id": int(company_id)},
    ).mappings().first()
    if not batch_exists:
        raise HTTPException(status_code=404, detail="Not Found")

    where_clauses = [
        "mr.company_id = :company_id",
        "mr.load_batch_id = :batch_id",
    ]
    params: dict[str, object] = {
        "company_id": int(company_id),
        "batch_id": int(batch_id),
        "limit": int(per_page),
        "offset": int((page - 1) * per_page),
    }

    normalized_status = str(status or "").strip().lower()
    if normalized_status in {"active", "rolled_back"}:
        where_clauses.append("LOWER(mr.status) = :status")
        params["status"] = normalized_status

    if product_id is not None:
        where_clauses.append("mr.product_id = :product_id")
        params["product_id"] = int(product_id)

    where_sql = " AND ".join(where_clauses)

    total_row = db.execute(
        text(
            f"""
            SELECT COUNT(1) AS total
            FROM capitados_monthly_records mr
            WHERE {where_sql}
            """
        ),
        params,
    ).mappings().first()
    total = int((total_row or {}).get("total") or 0)
    last_page = max(1, (total + per_page - 1) // per_page)

    rows = db.execute(
        text(
            f"""
            SELECT
                mr.id,
                mr.company_id,
                mr.product_id,
                mr.person_id,
                mr.contract_id,
                mr.coverage_month,
                mr.plan_version_id,
                mr.load_batch_id,
                mr.full_name,
                mr.sex,
                mr.age_reported,
                mr.price_base,
                mr.price_source,
                mr.age_surcharge_rule_id,
                mr.age_surcharge_percent,
                mr.age_surcharge_amount,
                mr.price_final,
                mr.status,
                mr.created_at,
                mr.updated_at,
                p.id AS person_ref_id,
                p.document_number AS person_document_number,
                p.full_name AS person_full_name,
                p.status AS person_status,
                rc.id AS residence_country_ref_id,
                rc.name AS residence_country_name,
                rc.iso2 AS residence_country_iso2,
                rc.iso3 AS residence_country_iso3,
                rep.id AS repatriation_country_ref_id,
                rep.name AS repatriation_country_name,
                rep.iso2 AS repatriation_country_iso2,
                rep.iso3 AS repatriation_country_iso3
            FROM capitados_monthly_records mr
            LEFT JOIN capitados_product_insureds p ON p.id = mr.person_id
            LEFT JOIN countries rc ON rc.id = mr.residence_country_id
            LEFT JOIN countries rep ON rep.id = mr.repatriation_country_id
            WHERE {where_sql}
            ORDER BY mr.id DESC
            LIMIT :limit OFFSET :offset
            """
        ),
        params,
    ).mappings().all()

    data = []
    for row in rows:
        data.append(
            {
                "id": int(row.get("id") or 0),
                "company_id": int(row.get("company_id") or 0),
                "product_id": row.get("product_id"),
                "person_id": row.get("person_id"),
                "contract_id": row.get("contract_id"),
                "coverage_month": row.get("coverage_month"),
                "plan_version_id": row.get("plan_version_id"),
                "load_batch_id": row.get("load_batch_id"),
                "full_name": row.get("full_name"),
                "sex": row.get("sex"),
                "age_reported": row.get("age_reported"),
                "price_base": row.get("price_base"),
                "price_source": row.get("price_source"),
                "age_surcharge_rule_id": row.get("age_surcharge_rule_id"),
                "age_surcharge_percent": row.get("age_surcharge_percent"),
                "age_surcharge_amount": row.get("age_surcharge_amount"),
                "price_final": row.get("price_final"),
                "status": row.get("status"),
                "created_at": row.get("created_at"),
                "updated_at": row.get("updated_at"),
                "person": {
                    "id": row.get("person_ref_id"),
                    "document_number": row.get("person_document_number"),
                    "full_name": row.get("person_full_name"),
                    "status": row.get("person_status"),
                }
                if row.get("person_ref_id") is not None
                else None,
                "residence_country": {
                    "id": row.get("residence_country_ref_id"),
                    "name": row.get("residence_country_name"),
                    "iso2": row.get("residence_country_iso2"),
                    "iso3": row.get("residence_country_iso3"),
                }
                if row.get("residence_country_ref_id") is not None
                else None,
                "repatriation_country": {
                    "id": row.get("repatriation_country_ref_id"),
                    "name": row.get("repatriation_country_name"),
                    "iso2": row.get("repatriation_country_iso2"),
                    "iso3": row.get("repatriation_country_iso3"),
                }
                if row.get("repatriation_country_ref_id") is not None
                else None,
                "can_rollback": _can_rollback_monthly_record(row),
            }
        )

    products_rows = db.execute(
        text(
            """
            SELECT DISTINCT p.id, p.name
            FROM capitados_monthly_records mr
            INNER JOIN products p ON p.id = mr.product_id
            WHERE mr.company_id = :company_id
              AND mr.load_batch_id = :batch_id
            ORDER BY p.id ASC
            """
        ),
        {"company_id": int(company_id), "batch_id": int(batch_id)},
    ).mappings().all()
    products = [
        {
            "id": int(row.get("id") or 0),
            "name": _decode_product_name_for_response(row.get("name")),
        }
        for row in products_rows
    ]

    return {
        "data": data,
        "meta": {
            "current_page": int(page),
            "last_page": int(last_page),
            "per_page": int(per_page),
            "total": int(total),
        },
        "products": products,
        "filters": {
            "status": normalized_status if normalized_status else "",
            "product_id": int(product_id) if product_id is not None else None,
        },
    }


@router.post("/{company_id}/capitated/batches/{batch_id:int}/rollback")
def batches_rollback(
    company_id: int,
    batch_id: int,
    request: Request,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
) -> dict:
    auth_payload = _require_admin(request=request, authorization=authorization, db=db)
    _require_company_exists(db=db, company_id=int(company_id))

    row = db.execute(
        text(
            """
            SELECT id, status, rolled_back_at
            FROM capitados_batch_logs
            WHERE id = :batch_id
              AND company_id = :company_id
            LIMIT 1
            """
        ),
        {"batch_id": int(batch_id), "company_id": int(company_id)},
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Not Found")

    if not _can_rollback_batch(row):
        raise HTTPException(status_code=422, detail={"message": "El lote no es elegible para rollback."})

    db.execute(
        text(
            """
            UPDATE capitados_monthly_records
            SET status = 'rolled_back',
                updated_at = NOW()
            WHERE company_id = :company_id
              AND load_batch_id = :batch_id
              AND LOWER(status) = 'active'
            """
        ),
        {"company_id": int(company_id), "batch_id": int(batch_id)},
    )

    db.execute(
        text(
            """
            UPDATE capitados_batch_logs
            SET status = 'rolled_back',
                rolled_back_at = NOW(),
                rolled_back_by_user_id = :rolled_back_by_user_id,
                updated_at = NOW()
            WHERE id = :batch_id
            """
        ),
        {
            "batch_id": int(batch_id),
            "rolled_back_by_user_id": int(auth_payload.get("id") or 0) or None,
        },
    )

    refreshed = db.execute(
        text(
            """
            SELECT
                b.id,
                b.company_id,
                b.coverage_month,
                b.source,
                b.source_file_id,
                b.original_filename,
                b.file_hash,
                b.created_by_user_id,
                b.status,
                b.processed_at,
                b.rolled_back_at,
                b.rolled_back_by_user_id,
                b.total_rows,
                b.total_applied,
                b.total_rejected,
                b.total_duplicated,
                b.total_incongruences,
                b.total_plan_errors,
                b.total_rolled_back,
                b.is_any_month_allowed,
                b.cutoff_day,
                b.error_summary,
                b.summary_json,
                b.created_at,
                b.updated_at,
                f.uuid AS source_file_uuid,
                u.id AS created_by_id,
                u.display_name AS created_by_display_name,
                u.first_name AS created_by_first_name,
                u.last_name AS created_by_last_name,
                u.email AS created_by_email
            FROM capitados_batch_logs b
            LEFT JOIN files f ON f.id = b.source_file_id
            LEFT JOIN users u ON u.id = b.created_by_user_id
            WHERE b.id = :batch_id
              AND b.company_id = :company_id
            LIMIT 1
            """
        ),
        {"batch_id": int(batch_id), "company_id": int(company_id)},
    ).mappings().first()

    db.commit()

    if not refreshed:
        raise HTTPException(status_code=500, detail="No se pudo actualizar el lote.")

    payload = _serialize_batch_row(refreshed)
    payload["can_rollback"] = False
    return {"batch": payload}


@router.post("/{company_id}/capitated/batches/{batch_id:int}/monthly-records/{record_id:int}/rollback")
def batches_monthly_record_rollback(
    company_id: int,
    batch_id: int,
    record_id: int,
    request: Request,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
) -> dict:
    _require_admin(request=request, authorization=authorization, db=db)
    _require_company_exists(db=db, company_id=int(company_id))

    row = db.execute(
        text(
            """
            SELECT id, status
            FROM capitados_monthly_records
            WHERE id = :record_id
              AND company_id = :company_id
              AND load_batch_id = :batch_id
            LIMIT 1
            """
        ),
        {
            "record_id": int(record_id),
            "company_id": int(company_id),
            "batch_id": int(batch_id),
        },
    ).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="Not Found")

    if not _can_rollback_monthly_record(row):
        raise HTTPException(status_code=422, detail={"message": "El registro no es elegible para rollback."})

    db.execute(
        text(
            """
            UPDATE capitados_monthly_records
            SET status = 'rolled_back',
                updated_at = NOW()
            WHERE id = :record_id
            """
        ),
        {"record_id": int(record_id)},
    )
    db.commit()

    return {"message": "Registro mensual revertido correctamente."}


@router.get("/{company_id}/capitated/batches")
def batches_index(
    company_id: int,
    request: Request,
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=15, ge=1, le=250),
    status: str | None = Query(default=None),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
) -> dict:
    _require_admin(request=request, authorization=authorization, db=db)
    _require_company_exists(db=db, company_id=int(company_id))

    where_clauses = ["b.company_id = :company_id"]
    params: dict[str, object] = {
        "company_id": int(company_id),
        "limit": int(per_page),
        "offset": int((page - 1) * per_page),
    }

    normalized_status = str(status or "").strip().lower()
    if normalized_status:
        where_clauses.append("LOWER(b.status) = :status")
        params["status"] = normalized_status

    where_sql = " AND ".join(where_clauses)

    total_row = db.execute(
        text(
            f"""
            SELECT COUNT(1) AS total
            FROM capitados_batch_logs b
            WHERE {where_sql}
            """
        ),
        params,
    ).mappings().first()
    total = int((total_row or {}).get("total") or 0)
    last_page = max(1, (total + per_page - 1) // per_page)

    rows = db.execute(
        text(
            f"""
            SELECT
                b.id,
                b.company_id,
                b.coverage_month,
                b.source,
                b.source_file_id,
                b.original_filename,
                b.file_hash,
                b.created_by_user_id,
                b.status,
                b.processed_at,
                b.rolled_back_at,
                b.rolled_back_by_user_id,
                b.total_rows,
                b.total_applied,
                b.total_rejected,
                b.total_duplicated,
                b.total_incongruences,
                b.total_plan_errors,
                b.total_rolled_back,
                b.is_any_month_allowed,
                b.cutoff_day,
                b.error_summary,
                b.summary_json,
                b.created_at,
                b.updated_at,
                f.uuid AS source_file_uuid
            FROM capitados_batch_logs b
            LEFT JOIN files f ON f.id = b.source_file_id
            WHERE {where_sql}
            ORDER BY b.id DESC
            LIMIT :limit OFFSET :offset
            """
        ),
        params,
    ).mappings().all()

    return {
        "data": [_serialize_batch_row(row) for row in rows],
        "meta": {
            "current_page": int(page),
            "last_page": int(last_page),
            "per_page": int(per_page),
            "total": int(total),
        },
    }


@router.get("/{company_id}/capitated/batches/template")
def batches_template(
    company_id: int,
    request: Request,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    _require_admin(request=request, authorization=authorization, db=db)
    _require_company_exists(db=db, company_id=int(company_id))

    rows = db.execute(
        text(
            """
            SELECT p.id, p.name
            FROM products p
            WHERE p.company_id = :company_id
              AND LOWER(p.status) = 'active'
              AND LOWER(p.product_type) = 'plan_capitado'
              AND EXISTS (
                  SELECT 1
                  FROM plan_versions pv
                  WHERE pv.product_id = p.id
                    AND LOWER(pv.status) = 'active'
              )
            ORDER BY p.id ASC
            """
        ),
        {"company_id": int(company_id)},
    ).mappings().all()

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: set[str] = set()
    headers = ["ID", "Nombre", "Residencia", "Nacionalidad", "Sexo", "Edad"]
    widths = {"A": 10, "B": 38, "C": 18, "D": 18, "E": 19, "F": 12}

    if not rows:
        sheet = workbook.create_sheet(title="Sin productos")
        sheet.append(headers)
        sheet.freeze_panes = "A2"
        sheet["A1"].font = sheet["B1"].font.copy(bold=True)
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
    else:
        for row in rows:
            product_name = _decode_product_name(row.get("name"))
            base_title = f"({int(row.get('id') or 0)}) {product_name}".strip()
            title = _excel_sheet_title(base_title, used_titles)

            sheet = workbook.create_sheet(title=title)
            sheet.append(headers)
            sheet.freeze_panes = "A2"
            sheet["A1"].font = sheet["B1"].font.copy(bold=True)
            for column, width in widths.items():
                sheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"capitados_estructura_company_{int(company_id)}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{company_id}/capitated/batches/upload")
async def batches_upload(
    company_id: int,
    request: Request,
    file: UploadFile = File(...),
    coverage_month: str | None = Form(default=None),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
) -> dict:
    auth_payload = _require_admin(request=request, authorization=authorization, db=db)
    _require_company_exists(db=db, company_id=int(company_id))

    permissions = {str(item or "").strip() for item in (auth_payload.get("permissions") or [])}
    can_create = "capitados.batch.create" in permissions
    can_create_any_month = "capitados.batch.create_any_month" in permissions
    if not can_create and not can_create_any_month:
        raise HTTPException(status_code=403, detail="No tiene permisos para crear batches de capitados.")

    filename = str(file.filename or "").strip()
    extension = Path(filename).suffix.lower()
    if extension not in {".xlsx", ".xls"}:
        raise HTTPException(status_code=422, detail="El archivo debe ser Excel (.xls o .xlsx).")

    normalized_month = _normalize_coverage_month(coverage_month)
    cutoff_day = 15

    if not can_create_any_month:
        now = datetime.utcnow().date()
        current_month = date(now.year, now.month, 1)
        if normalized_month != current_month:
            raise HTTPException(status_code=422, detail="Solo se permite cargar el mes en curso.")
        if now.day > cutoff_day:
            raise HTTPException(status_code=422, detail="La ventana de carga para el mes en curso ha expirado.")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=422, detail="El archivo subido no es valido.")

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="No fue posible leer el archivo Excel.") from exc

    total_rows = 0
    for sheet in workbook.worksheets:
        total_rows += max(0, int(sheet.max_row or 0) - 1)

    file_uuid = str(uuid4())
    relative_path = f"companies/{int(company_id)}/capitados/batches/{file_uuid}{extension}"
    disk = "public"

    storage = _storage_root() / disk
    absolute_path = (storage / relative_path).resolve()
    absolute_path.parent.mkdir(parents=True, exist_ok=True)
    absolute_path.write_bytes(file_bytes)

    file_hash = hashlib.sha1(file_bytes).hexdigest()
    mime_type = str(file.content_type or "").strip() or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    db.execute(
        text(
            """
            INSERT INTO files (uuid, disk, path, original_name, mime_type, size, uploaded_by, meta, created_at, updated_at)
            VALUES (:uuid, :disk, :path, :original_name, :mime_type, :size, :uploaded_by, :meta, NOW(), NOW())
            """
        ),
        {
            "uuid": file_uuid,
            "disk": disk,
            "path": relative_path,
            "original_name": filename,
            "mime_type": mime_type,
            "size": int(len(file_bytes)),
            "uploaded_by": int(auth_payload.get("id") or 0) or None,
            "meta": json.dumps({"context": "capitados_batch_source", "company_id": int(company_id)}),
        },
    )

    file_row = db.execute(
        text("SELECT LAST_INSERT_ID() AS id"),
    ).mappings().first()
    source_file_id = int((file_row or {}).get("id") or 0)

    db.execute(
        text(
            """
            INSERT INTO capitados_batch_logs (
                company_id,
                coverage_month,
                source,
                source_file_id,
                original_filename,
                file_hash,
                created_by_user_id,
                status,
                processed_at,
                total_rows,
                total_applied,
                total_rejected,
                total_duplicated,
                total_incongruences,
                total_plan_errors,
                total_rolled_back,
                is_any_month_allowed,
                cutoff_day,
                created_at,
                updated_at
            ) VALUES (
                :company_id,
                :coverage_month,
                'excel',
                :source_file_id,
                :original_filename,
                :file_hash,
                :created_by_user_id,
                'processed',
                NOW(),
                :total_rows,
                0,
                :total_rejected,
                0,
                0,
                0,
                0,
                :is_any_month_allowed,
                :cutoff_day,
                NOW(),
                NOW()
            )
            """
        ),
        {
            "company_id": int(company_id),
            "coverage_month": normalized_month.strftime("%Y-%m-01"),
            "source_file_id": source_file_id,
            "original_filename": filename,
            "file_hash": file_hash,
            "created_by_user_id": int(auth_payload.get("id") or 0) or None,
            "total_rows": int(total_rows),
            "total_rejected": int(total_rows),
            "is_any_month_allowed": 1 if can_create_any_month else 0,
            "cutoff_day": cutoff_day,
        },
    )

    batch_row = db.execute(
        text("SELECT LAST_INSERT_ID() AS id"),
    ).mappings().first()
    batch_id = int((batch_row or {}).get("id") or 0)

    row = db.execute(
        text(
            """
            SELECT
                b.id,
                b.company_id,
                b.coverage_month,
                b.source,
                b.source_file_id,
                b.original_filename,
                b.file_hash,
                b.created_by_user_id,
                b.status,
                b.processed_at,
                b.rolled_back_at,
                b.rolled_back_by_user_id,
                b.total_rows,
                b.total_applied,
                b.total_rejected,
                b.total_duplicated,
                b.total_incongruences,
                b.total_plan_errors,
                b.total_rolled_back,
                b.is_any_month_allowed,
                b.cutoff_day,
                b.error_summary,
                b.summary_json,
                b.created_at,
                b.updated_at,
                f.uuid AS source_file_uuid
            FROM capitados_batch_logs b
            LEFT JOIN files f ON f.id = b.source_file_id
            WHERE b.id = :batch_id
            LIMIT 1
            """
        ),
        {"batch_id": batch_id},
    ).mappings().first()

    db.commit()

    if not row:
        raise HTTPException(status_code=500, detail="No fue posible crear el batch.")

    return {"batch": _serialize_batch_row(row)}